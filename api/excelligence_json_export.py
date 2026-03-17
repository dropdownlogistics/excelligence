"""
Excelligence Registry → Explorer JSON Export
Produces a single JSON file matching the shape the explorer expects.

Usage:
    python excelligence_json_export.py <registry.xlsx> <output_path>

Example:
    python excelligence_json_export.py Excelligence_Registry_v011_50.xlsx api/excelligence.json

Output shape:
{
  "meta": { "version": "0.1.1", "entries": 50, "edges": 118, ... },
  "entries": [ { "id": "FRM-0001", "name": "XLOOKUP", ... } ],
  "edges": [ { "source": "FRM-0001", "target": "FRM-0002", "type": "LEADS_TO" } ],
  "tags": [ { "entry_id": "FRM-0001", "tag_name": "Data Modeling" } ],
  "aliases": [ { "entry_id": "FRM-0001", "alias": "lookup" } ]
}
"""

import sys
import os
import json
from openpyxl import load_workbook
from datetime import datetime


def export_json(registry_path, output_path):
    wb = load_workbook(registry_path)

    entries = []
    for row in wb["tbl_Entry"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            entries.append({
                "id": row[0],
                "name": row[1],
                "type": row[2],
                "tier": row[3],
                "what_it_does": row[4] or "",
                "intent": row[5] or "",
                "example": row[6] or "",
                "failure_modes": row[7] or "",
                "performance_notes": row[8] or "",
                "governance_notes": row[9] or "",
                "created_at": str(row[10] or ""),
                "updated_at": str(row[11] or ""),
            })

    edges = []
    for row in wb["tbl_Edges"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            edges.append({
                "source": row[0],
                "target": row[1],
                "type": row[2],
            })

    tags = []
    for row in wb["tbl_Tags"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            tags.append({
                "entry_id": row[0],
                "tag_id": row[1],
                "tag_name": row[2],
            })

    aliases = []
    for row in wb["tbl_Aliases"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            aliases.append({
                "entry_id": row[0],
                "alias": row[1],
            })

    # Type distribution
    type_dist = {}
    for e in entries:
        type_dist[e["type"]] = type_dist.get(e["type"], 0) + 1

    # Tier distribution
    tier_dist = {}
    for e in entries:
        tier_dist[e["tier"]] = tier_dist.get(e["tier"], 0) + 1

    output = {
        "meta": {
            "schema_version": "0.1.1",
            "entry_count": len(entries),
            "edge_count": len(edges),
            "alias_count": len(aliases),
            "tag_count": len(tags),
            "type_distribution": type_dist,
            "tier_distribution": tier_dist,
            "exported_at": datetime.now().isoformat(),
            "source": "Excelligence_Registry_v011_50.xlsx",
            "product": "Excelligence · Dropdown Logistics",
        },
        "entries": entries,
        "edges": edges,
        "tags": tags,
        "aliases": aliases,
    }

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Exported: {output_path}")
    print(f"  Entries: {len(entries)}")
    print(f"  Edges:   {len(edges)}")
    print(f"  Tags:    {len(tags)}")
    print(f"  Aliases: {len(aliases)}")
    print(f"  Size:    {os.path.getsize(output_path) / 1024:.1f} KB")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excelligence_json_export.py <registry.xlsx> <output_path>")
        sys.exit(1)
    export_json(sys.argv[1], sys.argv[2])
